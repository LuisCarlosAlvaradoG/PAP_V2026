from datetime import date, timedelta

import cvxpy as cp
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy.optimize import linprog, curve_fit

# ============================================================
#  CONFIGURACIÓN GENERAL
# ============================================================
ARCHIVO_DATOS = "Datos flujo de efectivo.xlsx"
ARCHIVO_SALIDA = "Flujo de efectivo.xlsx"
ARCHIVO_PAGOS = "Pagos pendientes.xlsx"
ARCHIVO_OPTIMIZACION = "Optimizacion de pagos.xlsx"

MESES = {
    1: "enero",
    2: "febrero",
    3: "marzo",
    4: "abril",
    5: "mayo",
    6: "junio",
    7: "julio",
    8: "agosto",
    9: "septiembre",
    10: "octubre",
    11: "noviembre",
    12: "diciembre",
}

# Logo oficial de FV Procesados, incrustado en base64 (fondo blanco removido)
# para que la app sea un solo archivo y no dependa de imágenes externas.
LOGO_BASE64 = "https://i.imgur.com/SNggzEh.png"

# ============================================================
#  SISTEMA DE DISEÑO — derivado del logo FV Procesados
# ============================================================
VERDE_OSCURO = "#3D5A28"
VERDE_MEDIO = "#5B7F3B"
VERDE_OLIVA = "#8A9A4E"
VERDE_CLARO = "#EEF3E3"
ROJO_FRESA = "#D6453D"
CREMA_FONDO = "#FBFAF9"
SIDEBAR_BG = "#24331A"
TINTA = "#22301A"


# ============================================================
#  FUNCIONES AUXILIARES
# ============================================================
def formato_semana(inicio, fin):
    if inicio.month == fin.month:
        return f"{inicio.day:02d}-{fin.day:02d} {MESES[fin.month]} {fin.year}"
    return f"{inicio.day:02d} {MESES[inicio.month]}-{fin.day:02d} {MESES[fin.month]} {fin.year}"


def formato_titulo(texto):
    return str(texto).strip().title()


def obtener_categorias(df):
    return sorted(df["categoría"].dropna().astype(str).str.strip().unique())


def obtener_conceptos(df):
    return sorted(df["descripción"].dropna().astype(str).str.strip().unique())


def obtener_tipo_por_categoria(categoria):
    if categoria in ["Ventas", "Préstamo"]:
        return "Ingreso"
    return "Egreso"


def buscar_similares(texto, opciones):
    texto = texto.upper().strip()
    palabras = texto.split()
    if texto == "":
        return []
    coincidencias = []
    for opcion in opciones:
        opcion_mayus = str(opcion).upper().strip()
        if texto in opcion_mayus:
            coincidencias.append(opcion)
        elif palabras and all(palabra in opcion_mayus for palabra in palabras):
            coincidencias.append(opcion)
        elif palabras and any(palabra in opcion_mayus for palabra in palabras):
            coincidencias.append(opcion)
    return list(dict.fromkeys(coincidencias))[:8]


def moneda(valor):
    signo = "-" if valor < 0 else ""
    return f"{signo}${abs(valor):,.2f}"


# ------------------------------------------------------------
#  Formato compartido para las hojas de Excel generadas
# ------------------------------------------------------------
def aplicar_formato_tabla(ws, filas_resaltadas=None, fila_flujo_neto=None):
    """Aplica el mismo lenguaje visual (bordes, encabezado verde,
    números con separador de miles) a cualquier hoja generada por la app."""
    filas_resaltadas = filas_resaltadas or []

    verde_encabezado = PatternFill(
        fill_type="solid", fgColor=VERDE_OSCURO.replace("#", "")
    )
    verde_claro_fill = PatternFill(
        fill_type="solid", fgColor=VERDE_CLARO.replace("#", "")
    )
    verde_flujo_fill = PatternFill(
        fill_type="solid", fgColor=VERDE_MEDIO.replace("#", "")
    )

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = verde_encabezado

    for row in ws.iter_rows():
        if row[0].value in filas_resaltadas:
            for cell in row:
                cell.font = Font(bold=True, color=VERDE_OSCURO.replace("#", ""))
                cell.fill = verde_claro_fill
        if fila_flujo_neto and row[0].value == fila_flujo_neto:
            for cell in row:
                cell.font = Font(bold=True, size=13, color="FFFFFF")
                cell.fill = verde_flujo_fill

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2


def generar_flujo(df):
    df = df.copy()
    df["inicio_semana"] = pd.to_datetime(df["inicio_semana"], dayfirst=True)
    df["tipo"] = df["tipo"].astype(str).str.strip()
    df["categoría"] = df["categoría"].astype(str).str.strip()
    df["descripción"] = df["descripción"].astype(str).str.strip()
    df["importe"] = pd.to_numeric(df["importe"], errors="coerce").fillna(0)
    df["importe_flujo"] = df["importe"]
    df.loc[df["tipo"] == "Egreso", "importe_flujo"] *= -1

    flujo = (
        df.groupby(["categoría", "inicio_semana"])["importe_flujo"]
        .sum()
        .unstack(fill_value=0)
        .sort_index()
    )

    total_ingresos = (
        df[df["tipo"] == "Ingreso"]
        .groupby("inicio_semana")["importe"]
        .sum()
        .reindex(flujo.columns, fill_value=0)
    )
    total_egresos = (
        df[df["tipo"] == "Egreso"]
        .groupby("inicio_semana")["importe"]
        .sum()
        .reindex(flujo.columns, fill_value=0)
    )
    flujo_neto = total_ingresos - total_egresos
    saldo_inicial = 0
    saldo = saldo_inicial + flujo_neto.cumsum()

    flujo.loc["TOTAL INGRESOS"] = total_ingresos
    flujo.loc["TOTAL EGRESOS"] = -total_egresos
    flujo.loc["FLUJO NETO"] = flujo_neto
    flujo.loc["SALDO"] = saldo

    nombres_semanas = (
        df[["inicio_semana", "semana"]]
        .drop_duplicates()
        .set_index("inicio_semana")["semana"]
        .to_dict()
    )
    flujo.columns = [nombres_semanas[col] for col in flujo.columns]
    flujo.index.name = "Categoría"
    flujo = flujo.round(2)

    with pd.ExcelWriter(ARCHIVO_SALIDA, engine="openpyxl") as writer:
        flujo.to_excel(writer, sheet_name="Flujo")
        ws = writer.sheets["Flujo"]
        aplicar_formato_tabla(
            ws,
            filas_resaltadas=["TOTAL INGRESOS", "TOTAL EGRESOS", "SALDO"],
            fila_flujo_neto="FLUJO NETO",
        )
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                cell.number_format = "#,##0.00"

    return flujo


def actualizar_fin():
    st.session_state.fin_semana = st.session_state.inicio_semana + timedelta(days=5)


# ------------------------------------------------------------
#  Pagos pendientes: carga y guardado
# ------------------------------------------------------------
COLUMNAS_PAGOS = [
    "concepto",
    "categoría",
    "tipo",
    "importe",
    "semana",
    "inicio_semana",
    "fecha_registro",
    "estatus",
    "fecha_pago",
]


def cargar_pagos():
    try:
        dfp = pd.read_excel(ARCHIVO_PAGOS)
        for col in COLUMNAS_PAGOS:
            if col not in dfp.columns:
                dfp[col] = ""
        dfp = dfp[COLUMNAS_PAGOS]
    except FileNotFoundError:
        dfp = pd.DataFrame(columns=COLUMNAS_PAGOS)
        guardar_pagos(dfp)
    return dfp


def guardar_pagos(dfp):
    with pd.ExcelWriter(ARCHIVO_PAGOS, engine="openpyxl") as writer:
        dfp.to_excel(writer, sheet_name="Pagos", index=False)
        ws = writer.sheets["Pagos"]
        aplicar_formato_tabla(ws, filas_resaltadas=[])
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"


# ============================================================
#  OPTIMIZACIÓN DE PAGOS
#  Implementación de los modelos descritos en el documento
#  "Modelo de Optimización F.E. FV-Procesados"
# ============================================================
def obtener_deudas_pendientes(dfp):
    """Agrupa la deuda pendiente por concepto/proveedor (D_i).

    Ordenado de MENOR a MAYOR a propósito: como el snowball paga las deudas
    más chicas primero, este orden deja la actividad de pago real visible
    desde el principio de las tablas de recomendación, en vez de escondida
    hasta el final.
    """
    pend = dfp[dfp["estatus"] == "Pendiente"].copy()
    pend["importe"] = pd.to_numeric(pend["importe"], errors="coerce").fillna(0)
    resumen = pend.groupby("concepto")["importe"].sum()
    return resumen[resumen > 0].sort_values(ascending=True)


def calcular_saldo_actual(df):
    """Saldo actual (C0) = ingresos - egresos de TODO el histórico registrado,
    sin filtrar nada -- a diferencia de la proyección, aquí sí debe contar
    cualquier movimiento ya registrado (incluida la semana en curso, o una
    inyección de capital de una sola vez), porque ya es saldo real, sin
    importar qué tan "representativo" sea de una semana típica.
    """
    df2 = df.copy()
    df2["importe"] = pd.to_numeric(df2["importe"], errors="coerce").fillna(0)
    ingresos = df2.loc[df2["tipo"] == "Ingreso", "importe"].sum()
    egresos = df2.loc[df2["tipo"] == "Egreso", "importe"].sum()
    return float(ingresos - egresos)


def obtener_proyeccion_ponderada(df, horizonte, ventana=8):
    """
    Proyección simple de ingresos y egresos semanales con media móvil
    PONDERADA, ventana de las últimas `ventana` semanas de historial (por
    defecto 8). A diferencia de un promedio simple, las semanas más
    recientes pesan más: con ventana=8 los pesos son 1,2,3,4,5,6,7,8 (la
    semana más vieja de las 8 pesa 1, la más reciente pesa 8), normalizados
    para sumar 1. Si hay menos semanas de historial que la ventana, se usan
    las que haya y los pesos se reescalan a esa cantidad.

    Antes de calcular, se excluye únicamente la semana en curso (todavía no
    termina, así que no es un dato completo).

    El promedio ponderado resultante (uno para ingresos, uno para egresos)
    se aplica de forma constante a cada una de las próximas `horizonte`
    semanas -- no hay estacionalidad ni otros ajustes adicionales.

    Devuelve dos listas de longitud `horizonte`:
        I : ingresos proyectados por semana (constante)
        E : egresos proyectados por semana (constante)
    """
    df = df.copy()
    df["importe"] = pd.to_numeric(df["importe"], errors="coerce").fillna(0)

    hoy = pd.Timestamp.today().normalize()
    fin_de_semana = df["inicio_semana"] + pd.Timedelta(days=6)
    df = df.loc[fin_de_semana < hoy].copy()

    if df.empty:
        return [0.0] * horizonte, [0.0] * horizonte

    resumen = df.pivot_table(
        index="inicio_semana", columns="tipo", values="importe",
        aggfunc="sum", fill_value=0,
    ).sort_index()

    for col in ["Ingreso", "Egreso"]:
        if col not in resumen.columns:
            resumen[col] = 0.0

    ingreso_hist = resumen["Ingreso"].values.astype(float)
    egreso_hist = resumen["Egreso"].values.astype(float)

    pesos_base = np.arange(1, ventana + 1, dtype=float)

    def wma(serie):
        ultimas = serie[-ventana:]
        n = len(ultimas)
        pesos = pesos_base[-n:]
        pesos = pesos / pesos.sum()
        return float(np.dot(ultimas, pesos))

    ingreso_prom = wma(ingreso_hist)
    egreso_prom = wma(egreso_hist)

    I = [ingreso_prom] * horizonte
    E = [egreso_prom] * horizonte
    return I, E


def modelo_a_pareto(D0, I, E, C0, L_rojo, T, n_candidatos=2001):
    """
    Modelo A - Frontera de Pareto / distancia al punto ideal (TOPSIS simplificado).

    Responde una sola pregunta -- "de la caja que tengo ahora mismo (C0),
    ¿cuánto le pago a la deuda pendiente (D0) EN TOTAL, de una sola vez?" --
    comparando el trade-off Caja-vs-Deuda como una frontera de Pareto, y
    escogiendo el punto más cercano al ideal.

    Usa la proyección (I, E, con media móvil ponderada de 8 semanas) para no dejar que el pago
    hoy hunda la caja proyectada por debajo del piso rojo dentro del
    horizonte T, pero SIEMPRE reporta dos cifras de caja por separado:
        - caja_resultante  : cómo queda la caja justo después de pagar hoy,
                              SIN proyección (C0 - pago).
        - caja_proyectada  : cómo se ve esa caja al final del horizonte T,
                              sumando el pago Y la proyección de ingresos y
                              egresos (C0 - pago + K), donde K es el flujo
                              neto proyectado acumulado a T semanas.

    Cómo funciona, paso a paso:
      1. Se arma una lista de montos de pago candidatos, de $0 hasta un tope
         máximo (el más restrictivo entre: la deuda total, y el monto que
         agotaría la caja PROYECTADA justo hasta el piso rojo/de emergencia
         -- nunca se considera un pago que cruce ese piso al final de T
         semanas).
      2. Para cada candidato, se calcula la Caja proyectada resultante y la
         Deuda restante.
      3. El "punto ideal" es el mejor valor observado de cada criterio por
         separado: la Caja proyectada más alta posible (pagar $0) y la
         Deuda más baja posible (pagar el tope máximo) -- un punto que en
         la práctica ningún candidato alcanza en ambos criterios a la vez.
      4. Se mide qué tan lejos está cada candidato de ese punto ideal
         (distancia euclidiana en el plano Caja-Deuda), y se elige el
         candidato con la menor distancia.

    Como la Caja y la Deuda se mueven exactamente $1 por cada $1 pagado, el
    punto más cercano al ideal siempre resulta ser la MITAD del tope máximo
    considerado -- ni el extremo de no pagar nada, ni el extremo de pagar
    todo lo posible.

    Una vez decidido el monto total óptimo, se reparte entre las deudas
    individuales con snowball estricto (la más chica primero).

    Retorna: pago_por_concepto (array), diagnostico (dict con pago_total,
    caja_resultante, caja_proyectada, deuda_resultante,
    pago_max_considerado, K), exito (bool).
    """
    D0 = np.asarray(D0, dtype=float)
    n = len(D0)
    D0_total = D0.sum()
    I_arr = np.asarray(I, dtype=float)
    E_arr = np.asarray(E, dtype=float)

    K = float(np.sum(I_arr[:T] - E_arr[:T]))  # flujo neto proyectado acumulado a T semanas

    pago_max = max(0.0, min(D0_total, C0 + K - L_rojo))

    if pago_max <= 0:
        # Ni pagando $0 se respeta el piso rojo en la caja proyectada --
        # infactible, hace falta revisar el saldo actual, la proyección o
        # el umbral rojo antes de pedir pagar deuda.
        return np.zeros(n), {}, False

    candidatos = np.linspace(0.0, pago_max, n_candidatos)
    caja_proyectada_cand = C0 - candidatos + K
    deuda_cand = D0_total - candidatos

    caja_ideal = caja_proyectada_cand.max()
    deuda_ideal = deuda_cand.min()
    distancias = np.sqrt(
        (deuda_cand - deuda_ideal) ** 2 + (caja_proyectada_cand - caja_ideal) ** 2
    )
    idx_opt = int(np.argmin(distancias))
    pago_optimo = float(candidatos[idx_opt])

    # Snowball: reparte el monto total decidido entre las deudas individuales
    D_restante = D0.copy()
    orden = np.argsort(D_restante)
    pago_por_concepto = np.zeros(n)
    restante = pago_optimo
    for i in orden:
        if restante <= 0:
            break
        abono = min(D_restante[i], restante)
        pago_por_concepto[i] = abono
        restante -= abono

    diagnostico = dict(
        pago_total=pago_optimo,
        caja_resultante=float(C0 - pago_optimo),
        caja_proyectada=float(caja_proyectada_cand[idx_opt]),
        deuda_resultante=float(deuda_cand[idx_opt]),
        pago_max_considerado=pago_max,
        distancia=float(distancias[idx_opt]),
        K=K,
    )
    return pago_por_concepto, diagnostico, True


def modelo2_resiliencia(I, E, C0, T, R, F, s, E_max=None, L_min=None, D0=None):
    """
    Modelo 2 - Máxima resiliencia financiera.

    NOTA: se conserva SOLO como referencia académica -- ya no aparece en el
    selector de la interfaz ni en la comparación interactiva de modelos. Por
    diseño, nunca paga ni un peso de deuda (solo reprograma egresos), así que
    no aporta al objetivo de "ver al modelo decidir cuánto pagar" que motivó
    esta versión del módulo. Queda documentada aquí por si se necesita
    referenciarla o retomarla más adelante.

    Extremo CONSERVADOR del espectro: por diseño nunca toca D0 ni paga un peso de
    deuda (solo reprograma egresos flexibles dentro de una ventana +-s semanas,
    respetando los egresos rígidos). Maximiza el piso de caja alcanzable.

    L_min/L_survival ya NO son restricción dura (solo referencia informativa): con
    flujos donde egresos > ingresos, el piso máximo alcanzable suele ser inferior a
    cualquier L_min positivo, y exigirlo como obligatorio vuelve el problema
    infeasible sin importar cómo se reacomoden los pagos.

    Se reporta también la "resiliencia ganada" (m_opt - m_base): cuánto mejora el
    piso gracias a la reprogramación vs. no reprogramar nada. Matemáticamente esta
    ganancia es $0 si la semana crítica cae en la ÚLTIMA semana del horizonte (la
    caja acumulada al final del horizonte es invariante a cómo se reordenen los
    pagos) -- no es un error, es una propiedad esperada del modelo.
    """
    I = np.asarray(I, dtype=float)
    E = np.asarray(E, dtype=float)
    R = np.asarray(R, dtype=float)
    F = np.asarray(F, dtype=float)
    if not (len(I) == len(E) == len(R) == len(F) == T):
        raise ValueError("I, E, R, F deben tener longitud T")
    if np.any(F < 0) or np.any(R < 0):
        raise ValueError("R y F deben ser no negativos")

    C_base = np.zeros(T + 1)
    C_base[0] = C0
    for t in range(1, T + 1):
        C_base[t] = C_base[t - 1] + I[t - 1] - E[t - 1]
    m_base = float(np.min(C_base[1:]))

    if np.all(F == 0):
        return dict(
            m_opt=m_base, m_base=m_base, ganancia=0.0,
            C_opt=C_base.tolist(), E_opt=E.tolist(), x_opt={}, exito=True,
            t_crit=int(np.argmin(C_base[1:])) + 1,
            K_star=(np.max(E) / np.sum(E)) if np.sum(E) > 0 else np.nan
        )

    pairs, idx_map, n_x = [], {}, 0
    for tau in range(T):
        t_start, t_end = max(0, tau - s), min(T, tau + s + 1)
        for t in range(t_start, t_end):
            pairs.append((tau, t))
            idx_map[(tau, t)] = n_x
            n_x += 1

    n_vars = n_x + 1
    idx_m = n_x
    c = np.zeros(n_vars)
    c[idx_m] = -1.0

    A_eq, b_eq = [], []
    for tau in range(T):
        row = np.zeros(n_vars)
        t_start, t_end = max(0, tau - s), min(T, tau + s + 1)
        for t in range(t_start, t_end):
            row[idx_map[(tau, t)]] = 1.0
        A_eq.append(row)
        b_eq.append(F[tau])
    A_eq, b_eq = np.array(A_eq), np.array(b_eq)

    A_ub, b_ub = [], []
    cum_I, cum_R = np.cumsum(I), np.cumsum(R)
    for t_idx in range(1, T + 1):
        row = np.zeros(n_vars)
        for k in range(t_idx):
            tau_start, tau_end = max(0, k - s), min(T, k + s + 1)
            for tau in range(tau_start, tau_end):
                if (tau, k) in idx_map:
                    row[idx_map[(tau, k)]] += 1.0
        row[idx_m] = 1.0
        A_ub.append(row)
        b_ub.append(C0 + cum_I[t_idx - 1] - cum_R[t_idx - 1])

    if E_max is not None:
        E_max = np.asarray(E_max, dtype=float)
        for t in range(T):
            if np.isfinite(E_max[t]):
                row = np.zeros(n_vars)
                tau_start, tau_end = max(0, t - s), min(T, t + s + 1)
                for tau in range(tau_start, tau_end):
                    row[idx_map[(tau, t)]] = 1.0
                A_ub.append(row)
                b_ub.append(E_max[t] - R[t])

    A_ub = np.array(A_ub) if A_ub else None
    b_ub = np.array(b_ub) if b_ub else None
    bounds = [(0, None)] * n_x + [(None, None)]

    res = linprog(c, A_ub=A_ub, b_ub=b_ub, A_eq=A_eq, b_eq=b_eq, bounds=bounds, method="highs")

    if not res.success:
        return dict(m_opt=np.nan, m_base=m_base, ganancia=np.nan, C_opt=[], E_opt=[],
                    x_opt={}, exito=False, t_crit=None, K_star=np.nan)

    x_sol = res.x[:n_x]
    m_opt = res.x[idx_m]
    x_opt = {pair: x_sol[idx_map[pair]] for pair in pairs}

    E_opt = np.copy(R)
    for (tau, t), val in x_opt.items():
        E_opt[t] += val

    C_opt = np.zeros(T + 1)
    C_opt[0] = C0
    for t in range(1, T + 1):
        C_opt[t] = C_opt[t - 1] + I[t - 1] - E_opt[t - 1]

    t_crit = int(np.argmin(C_opt[1:])) + 1
    sum_E_opt = np.sum(E_opt)
    K_star = np.max(E_opt) / sum_E_opt if sum_E_opt > 0 else np.nan

    return dict(m_opt=m_opt, m_base=m_base, ganancia=float(m_opt - m_base),
                C_opt=C_opt.tolist(), E_opt=E_opt.tolist(), x_opt=x_opt,
                exito=True, t_crit=t_crit, K_star=K_star)


def construir_tabla_lista_pagos(pago_por_concepto, conceptos):
    """Tabla de recomendación SIMPLE: una lista de "Concepto | Pago Recomendado",
    sin desglose semanal -- ambos modelos (A y B) ahora responden "cuánto pagarle
    a cada quien", no "cuándo, semana por semana". Mantiene el orden de menor a
    mayor deuda (mismo criterio usado en toda la app) para que sea fácil ver de
    un vistazo cuáles se alcanzan a liquidar por completo."""
    filas = []
    for concepto, pago in zip(conceptos, pago_por_concepto):
        filas.append({
            "Concepto": concepto,
            "Pago Recomendado ($)": round(float(pago), 2),
        })
    return pd.DataFrame(filas)


def clasificar_caja(valor, verde, amarillo, rojo):
    """Clasifica un nivel de caja en el semáforo de 3 zonas y devuelve
    (etiqueta, color, cuánto falta/sobra hasta el siguiente umbral)."""
    if valor >= verde:
        return "🟢 Verde", "green", valor - verde
    elif valor >= amarillo:
        return "🟡 Amarillo", "orange", valor - amarillo
    elif valor >= rojo:
        return "🔴 Rojo", "red", valor - rojo
    else:
        return "⚫ Crítico (bajo el piso de emergencia)", "black", valor - rojo


def construir_tabla_modelo2_resiliencia(res, L_min=None, L_survival=None):
    """Tabla de recomendación para Modelo 2 (Resiliencia): no reparte pagos por
    concepto, solo muestra la caja y el egreso reprogramado semana a semana, más
    las métricas de piso/resiliencia ganada.

    La columna "Semana" se guarda como texto (combina números de semana con
    etiquetas de métricas), y las columnas numéricas usan NaN (no "") para las
    celdas en blanco -- una columna de tipo mixto (número + string) rompe la
    serialización Arrow que usa Streamlit para mostrar tablas.
    """
    if not res["exito"]:
        return pd.DataFrame({"Error": ["Optimización fallida"]})
    T_mas_1 = len(res["C_opt"])
    df_tabla = pd.DataFrame({
        "Semana": [str(t) for t in range(T_mas_1)],
        "Caja": res["C_opt"],
        "Egreso_reprogramado": [res["E_opt"][t - 1] if t > 0 else np.nan for t in range(T_mas_1)],
    })
    filas_metricas = [
        ("---", np.nan),
        ("Piso SIN reprogramar (m_base)", res["m_base"]),
        ("Piso CON reprogramación (m*)", res["m_opt"]),
        ("Resiliencia ganada (m* - m_base)", res["ganancia"]),
        ("Semana crítica", res["t_crit"]),
        ("Índice K*", res["K_star"]),
    ]
    if L_min is not None:
        filas_metricas.append(("L_min (colchón cómodo, referencia)", L_min))
    if L_survival is not None:
        filas_metricas.append(("L_survival (piso de emergencia, referencia)", L_survival))
    metricas = pd.DataFrame({
        "Semana": [f[0] for f in filas_metricas],
        "Caja": [f[1] for f in filas_metricas],
        "Egreso_reprogramado": [np.nan] * len(filas_metricas),
    })
    return pd.concat([df_tabla, metricas], ignore_index=True)


# ============================================================
#  PÁGINA + SISTEMA DE DISEÑO
# ============================================================
st.set_page_config(
    page_title="FV Procesados | Sistema Financiero",
    page_icon=None,
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    f"""
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,400;9..144,600;9..144,700&family=IBM+Plex+Sans:wght@400;500;600;700&family=IBM+Plex+Mono:wght@500;600&display=swap" rel="stylesheet">

<style>
    html, body, [class*="css"] {{
        font-family: 'IBM Plex Sans', sans-serif;
        color: {TINTA};
    }}
    .stApp {{
        background-color: {CREMA_FONDO};
    }}
    #MainMenu, header[data-testid="stHeader"] {{
        background-color: transparent;
    }}

    /* -------- Tipografía de títulos -------- */
    h1, h2, h3 {{
        font-family: 'Fraunces', serif;
        color: {TINTA} !important;
        letter-spacing: -0.01em;
    }}

    /* -------- La regla-firma (línea del logo) -------- */
    .regla-marca {{
        height: 2px;
        background-color: {VERDE_OLIVA};
        width: 100%;
        margin: 6px 0 22px 0;
        position: relative;
    }}
    .regla-marca::before {{
        content: "";
        position: absolute;
        left: 0;
        top: -3px;
        width: 8px;
        height: 8px;
        border-radius: 50%;
        background-color: {ROJO_FRESA};
    }}

    /* ================= SIDEBAR ================= */
    section[data-testid="stSidebar"] {{
        background-color: {SIDEBAR_BG};
        border-right: none;
    }}
    section[data-testid="stSidebar"] * {{
        color: {VERDE_CLARO};
    }}
    .marca-wordmark {{
        padding: 26px 24px 0 24px;
        text-align: center;
    }}
    .marca-wordmark .logo-sidebar {{
        width: 100%;
        max-width: 168px;
    }}
    .marca-wordmark .regla-marca {{
        margin: 18px 0 24px 0;
    }}

    .nav-eyebrow {{
        color: {VERDE_OLIVA};
        text-transform: uppercase;
        font-size: 11px;
        letter-spacing: 2px;
        font-weight: 600;
        padding: 4px 20px 8px 20px;
    }}

    section[data-testid="stSidebar"] .stButton > button {{
        background-color: transparent;
        border: none;
        border-left: 3px solid transparent;
        border-radius: 0;
        text-align: left;
        justify-content: flex-start;
        width: 100%;
        padding: 10px 20px;
        font-weight: 500;
        font-size: 15px;
        color: {VERDE_CLARO};
    }}
    section[data-testid="stSidebar"] .stButton > button:hover {{
        background-color: rgba(255,255,255,0.07);
        color: #FFFFFF;
        border-left: 3px solid {VERDE_OLIVA};
    }}
    section[data-testid="stSidebar"] .stButton > button:focus:not(:active) {{
        color: #FFFFFF;
    }}

    .nav-item-activo {{
        border-left: 3px solid {ROJO_FRESA};
        background-color: rgba(255,255,255,0.09);
        padding: 10px 17px;
        font-weight: 700;
        font-size: 15px;
        color: #FFFFFF;
        margin-bottom: 1px;
    }}
    .nav-item-inactivo-sep {{
        height: 10px;
    }}

    .nav-sub-activo {{
        border-left: 3px solid {ROJO_FRESA};
        background-color: rgba(255,255,255,0.06);
        padding: 10px 17px 10px 31px;
        font-weight: 600;
        font-size: 14px;
        color: #FFFFFF;
        margin-top: 1px !important;
        margin-bottom: 6px !important;
        display: block !important;
    }}
    section[data-testid="stSidebar"] .nav-sub .stButton > button {{
        padding-left: 34px;
        font-size: 14px;
        font-weight: 400;
        margin-top: 4px !important;
        margin-bottom: 4px !important;
    }}

    /* ================= CONTENIDO PRINCIPAL ================= */
    .encabezado-modulo {{
        margin-bottom: 4px;
    }}
    .encabezado-modulo .eyebrow {{
        color: {VERDE_OLIVA};
        text-transform: uppercase;
        letter-spacing: 2px;
        font-size: 12px;
        font-weight: 600;
    }}
    .encabezado-modulo h1 {{
        margin: 4px 0 0 0 !important;
        font-size: 30px !important;
    }}

    /* -------- Tarjetas KPI -------- */
    .kpi-tarjeta {{
        background-color: #FFFFFF;
        border: 1px solid rgba(0,0,0,0.06);
        border-left: 4px solid {VERDE_MEDIO};
        border-radius: 6px;
        padding: 16px 18px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }}
    .kpi-tarjeta .kpi-label {{
        text-transform: uppercase;
        font-size: 11px;
        letter-spacing: 1.5px;
        font-weight: 600;
        color: {VERDE_OLIVA};
    }}
    .kpi-tarjeta .kpi-valor {{
        font-family: 'IBM Plex Mono', monospace;
        font-size: 24px;
        font-weight: 600;
        color: {TINTA};
        margin-top: 4px;
    }}
    .kpi-egreso {{ border-left-color: {ROJO_FRESA}; }}
    .kpi-saldo {{ border-left-color: {VERDE_OSCURO}; }}

    /* -------- Contenedores tipo tarjeta (formularios, tabla) -------- */
    div[data-testid="stVerticalBlockBorderWrapper"] {{
        background-color: #FFFFFF;
        border-color: rgba(0,0,0,0.08) !important;
        border-radius: 8px !important;
    }}
    div[data-testid="stVerticalBlockBorderWrapper"] > div {{
        border-radius: 8px !important;
    }}

    /* -------- Pantalla de bienvenida: solo el logo, centrado -------- */
    .bienvenida-hero {{
        text-align: center;
        display: flex;
        align-items: center;
        justify-content: center;
        min-height: 70vh;
    }}
    .bienvenida-hero .logo-hero {{
        width: 100%;
        max-width: 420px;
    }}

    /* -------- Botones de la app -------- */
    .stButton > button {{
        border-radius: 6px;
        font-weight: 600;
    }}
    div[data-testid="stAppViewContainer"] .stButton > button {{
        background-color: {VERDE_MEDIO};
        color: white;
        border: none;
    }}
    div[data-testid="stAppViewContainer"] .stButton > button:hover {{
        background-color: {VERDE_OSCURO};
        color: white;
    }}
    .stDownloadButton > button {{
        background-color: {ROJO_FRESA};
        color: white;
        border: none;
        border-radius: 6px;
        font-weight: 700;
    }}
    .stDownloadButton > button:hover {{
        background-color: #B8342D;
        color: white;
    }}

    div[data-testid="stAlert"] {{
        border-radius: 8px;
    }}
</style>
""",
    unsafe_allow_html=True,
)

# ============================================================
#  ESTADO DE NAVEGACIÓN
# ============================================================
if "modulo_activo" not in st.session_state:
    st.session_state.modulo_activo = None
if "seccion_flujo" not in st.session_state:
    st.session_state.seccion_flujo = None
if "seccion_pagos" not in st.session_state:
    st.session_state.seccion_pagos = None

if "inicio_semana" not in st.session_state:
    st.session_state.inicio_semana = date.today()
if "fin_semana" not in st.session_state:
    st.session_state.fin_semana = date.today() + timedelta(days=5)
if "concepto_seleccionado" not in st.session_state:
    st.session_state.concepto_seleccionado = None
if "categoria_seleccionada" not in st.session_state:
    st.session_state.categoria_seleccionada = None

# Estado para el sub-formulario "Agregar deuda"
if "deuda_inicio_semana" not in st.session_state:
    st.session_state.deuda_inicio_semana = date.today()
if "deuda_fin_semana" not in st.session_state:
    st.session_state.deuda_fin_semana = date.today() + timedelta(days=5)


def ir_a(modulo, seccion=None):
    st.session_state.modulo_activo = modulo
    if modulo == "flujo":
        st.session_state.seccion_flujo = seccion
    elif modulo == "pagos":
        st.session_state.seccion_pagos = seccion
    st.rerun()


def actualizar_fin_deuda():
    st.session_state.deuda_fin_semana = (
        st.session_state.deuda_inicio_semana + timedelta(days=5)
    )


# ============================================================
#  SIDEBAR — NAVEGACIÓN
# ============================================================
with st.sidebar:
    st.markdown(
        f"""
    <div class="marca-wordmark">
        <img src="{LOGO_BASE64}" class="logo-sidebar" alt="FV Procesados">
        <div class="regla-marca"></div>
    </div>
    """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="nav-eyebrow">Módulos</div>', unsafe_allow_html=True)

    # ---- Módulo: Flujo de Efectivo ----
    if st.session_state.modulo_activo == "flujo":
        st.markdown(
            '<div class="nav-item-activo">Flujo de Efectivo</div>',
            unsafe_allow_html=True,
        )

        st.markdown('<div class="nav-sub">', unsafe_allow_html=True)
        secciones = [
            ("ver", "Ver flujo de efectivo"),
            ("agregar", "Agregar movimiento"),
            ("eliminar", "Eliminar movimiento"),
        ]
        for clave, etiqueta in secciones:
            if st.session_state.seccion_flujo == clave:
                st.markdown(
                    f'<div class="nav-sub-activo">{etiqueta}</div>',
                    unsafe_allow_html=True,
                )
            else:
                if st.button(etiqueta, key=f"nav_{clave}", use_container_width=True):
                    ir_a("flujo", clave)
        st.markdown("</div>", unsafe_allow_html=True)
    else:
        if st.button(
            "Flujo de Efectivo", key="nav_modulo_flujo", use_container_width=True
        ):
            ir_a("flujo", "ver")

    st.markdown('<div class="nav-item-inactivo-sep"></div>', unsafe_allow_html=True)

    # ---- Módulo: Pagos Pendientes ----
    if st.session_state.modulo_activo == "pagos":
        st.markdown(
            '<div class="nav-item-activo">Pagos Pendientes</div>',
            unsafe_allow_html=True,
        )

        st.markdown('<div class="nav-sub">', unsafe_allow_html=True)
        secciones_pagos = [
            ("agregar", "Agregar deuda"),
            ("pagar", "Pagar un pago pendiente"),
            ("optimizar", "Optimización de pagos"),
        ]
        for clave, etiqueta in secciones_pagos:
            if st.session_state.seccion_pagos == clave:
                st.markdown(
                    f'<div class="nav-sub-activo">{etiqueta}</div>',
                    unsafe_allow_html=True,
                )
            else:
                if st.button(
                    etiqueta, key=f"nav_pagos_{clave}", use_container_width=True
                ):
                    ir_a("pagos", clave)
        st.markdown("</div>", unsafe_allow_html=True)
    else:
        if st.button(
            "Pagos Pendientes", key="nav_modulo_pagos", use_container_width=True
        ):
            ir_a("pagos", "agregar")

# ============================================================
#  CARGA DE DATOS
# ============================================================
df = pd.read_excel(ARCHIVO_DATOS)
df["inicio_semana"] = pd.to_datetime(df["inicio_semana"], dayfirst=True)
df["tipo"] = df["tipo"].astype(str).str.strip()
df["categoría"] = df["categoría"].astype(str).str.strip()
df["descripción"] = df["descripción"].astype(str).str.strip()

TITULOS_SECCION = {
    "ver": "Ver flujo de efectivo",
    "agregar": "Agregar movimiento",
    "eliminar": "Eliminar movimiento",
}

TITULOS_SECCION_PAGOS = {
    "agregar": "Agregar deuda",
    "pagar": "Pagar un pago pendiente",
    "optimizar": "Optimización de pagos",
}

# ============================================================
#  MÓDULO 1: FLUJO DE EFECTIVO
# ============================================================
if st.session_state.modulo_activo == "flujo":
    seccion = st.session_state.seccion_flujo

    st.markdown(
        f"""
    <div class="encabezado-modulo">
        <div class="eyebrow">Flujo de Efectivo</div>
        <h1>{TITULOS_SECCION[seccion]}</h1>
    </div>
    <div class="regla-marca"></div>
    """,
        unsafe_allow_html=True,
    )

    # ---------------- VER FLUJO DE EFECTIVO ----------------
    if seccion == "ver":
        df_actual = pd.read_excel(ARCHIVO_DATOS)
        df_actual["tipo"] = df_actual["tipo"].astype(str).str.strip()
        df_actual["importe"] = pd.to_numeric(
            df_actual["importe"], errors="coerce"
        ).fillna(0)

        flujo = generar_flujo(df_actual)

        total_ingresos = df_actual.loc[df_actual["tipo"] == "Ingreso", "importe"].sum()
        total_egresos = df_actual.loc[df_actual["tipo"] == "Egreso", "importe"].sum()
        flujo_neto = total_ingresos - total_egresos
        saldo_actual = flujo.loc["SALDO"].iloc[-1] if flujo.shape[1] > 0 else 0

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(
                f"""
            <div class="kpi-tarjeta">
                <div class="kpi-label">Ingresos totales</div>
                <div class="kpi-valor">{moneda(total_ingresos)}</div>
            </div>""",
                unsafe_allow_html=True,
            )
        with c2:
            st.markdown(
                f"""
            <div class="kpi-tarjeta kpi-egreso">
                <div class="kpi-label">Egresos totales</div>
                <div class="kpi-valor">{moneda(total_egresos)}</div>
            </div>""",
                unsafe_allow_html=True,
            )
        with c3:
            st.markdown(
                f"""
            <div class="kpi-tarjeta">
                <div class="kpi-label">Flujo neto</div>
                <div class="kpi-valor">{moneda(flujo_neto)}</div>
            </div>""",
                unsafe_allow_html=True,
            )
        with c4:
            st.markdown(
                f"""
            <div class="kpi-tarjeta kpi-saldo">
                <div class="kpi-label">Saldo actual</div>
                <div class="kpi-valor">{moneda(saldo_actual)}</div>
            </div>""",
                unsafe_allow_html=True,
            )

        st.write("")
        with st.container(border=True):
            st.dataframe(flujo, use_container_width=True)

        with open(ARCHIVO_SALIDA, "rb") as archivo:
            st.download_button(
                label="Descargar Excel",
                data=archivo,
                file_name=ARCHIVO_SALIDA,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # ---------------- AGREGAR MOVIMIENTO ----------------
    elif seccion == "agregar":
        with st.container(border=True):
            col1, col2 = st.columns(2)
            with col1:
                inicio = st.date_input(
                    "Inicio de semana", key="inicio_semana", on_change=actualizar_fin
                )
            with col2:
                fin = st.date_input("Fin de semana", key="fin_semana")

            semana = formato_semana(inicio, fin)
            st.info(f"Semana seleccionada: {semana}")

            importe = st.number_input("Importe", min_value=0.0, step=100.0)

            st.markdown('<div class="regla-marca"></div>', unsafe_allow_html=True)

            conceptos_existentes = obtener_conceptos(df)
            categorias_existentes = obtener_categorias(df)

            modo_concepto = st.radio(
                "Concepto",
                ["Usar concepto existente", "Agregar concepto nuevo"],
                horizontal=True,
            )

            concepto = ""
            categoria = ""
            tipo = ""

            if modo_concepto == "Usar concepto existente":
                concepto = st.selectbox(
                    "Selecciona el concepto", [""] + conceptos_existentes
                )
                if concepto != "":
                    datos_concepto = df[
                        df["descripción"].astype(str).str.strip() == concepto
                    ].iloc[0]
                    categoria = datos_concepto["categoría"]
                    tipo = datos_concepto["tipo"]
                    st.text_input("Categoría", value=categoria, disabled=True)
                    st.text_input("Tipo", value=tipo, disabled=True)
            else:
                concepto_escrito = st.text_input(
                    "Nuevo concepto", placeholder="Escribe el concepto"
                ).strip()

                if st.session_state.concepto_seleccionado is not None:
                    concepto = st.session_state.concepto_seleccionado
                    datos_concepto = df[
                        df["descripción"].astype(str).str.strip() == concepto
                    ].iloc[0]
                    categoria = datos_concepto["categoría"]
                    tipo = datos_concepto["tipo"]
                    st.text_input(
                        "Concepto seleccionado", value=concepto, disabled=True
                    )
                    st.text_input("Categoría", value=categoria, disabled=True)
                    st.text_input("Tipo", value=tipo, disabled=True)
                else:
                    concepto = formato_titulo(concepto_escrito)
                    coincidencias_concepto = buscar_similares(
                        concepto_escrito, conceptos_existentes
                    )
                    if concepto_escrito and coincidencias_concepto:
                        st.caption("Coincidencias encontradas:")
                        columnas = st.columns(4)
                        for i, sugerencia in enumerate(coincidencias_concepto):
                            with columnas[i % 4]:
                                if st.button(sugerencia, key=f"concepto_{i}"):
                                    st.session_state.concepto_seleccionado = sugerencia
                                    st.rerun()

                    modo_categoria = st.radio(
                        "Categoría",
                        ["Usar categoría existente", "Agregar categoría nueva"],
                        horizontal=True,
                    )

                    if modo_categoria == "Usar categoría existente":
                        categoria = st.selectbox(
                            "Selecciona la categoría", [""] + categorias_existentes
                        )
                        if categoria != "":
                            tipo = obtener_tipo_por_categoria(categoria)
                            st.text_input("Tipo", value=tipo, disabled=True)
                    else:
                        categoria_escrita = st.text_input(
                            "Nueva categoría", placeholder="Escribe la categoría"
                        ).strip()

                        if st.session_state.categoria_seleccionada is not None:
                            categoria = st.session_state.categoria_seleccionada
                            tipo = obtener_tipo_por_categoria(categoria)
                            st.text_input(
                                "Categoría seleccionada", value=categoria, disabled=True
                            )
                            st.text_input("Tipo", value=tipo, disabled=True)
                        else:
                            categoria = formato_titulo(categoria_escrita)
                            coincidencias_categoria = buscar_similares(
                                categoria_escrita, categorias_existentes
                            )
                            if categoria_escrita and coincidencias_categoria:
                                st.caption("Coincidencias encontradas:")
                                columnas_cat = st.columns(4)
                                for i, sugerencia in enumerate(coincidencias_categoria):
                                    with columnas_cat[i % 4]:
                                        if st.button(sugerencia, key=f"categoria_{i}"):
                                            st.session_state.categoria_seleccionada = (
                                                sugerencia
                                            )
                                            st.rerun()
                            if categoria_escrita:
                                tipo = st.selectbox(
                                    "Tipo", ["Ingreso", "Egreso"], index=1
                                )

            if st.button("Guardar movimiento"):
                if concepto == "":
                    st.error("Debes seleccionar o escribir un concepto.")
                elif importe <= 0:
                    st.error("El importe debe ser mayor a 0.")
                elif categoria == "":
                    st.error("Debes seleccionar o escribir una categoría.")
                elif tipo == "":
                    st.error("Debes seleccionar o confirmar el tipo.")
                elif fin < inicio:
                    st.error("La fecha final no puede ser anterior a la fecha inicial.")
                else:
                    nuevo = pd.DataFrame(
                        [
                            {
                                "inicio_semana": pd.to_datetime(inicio),
                                "semana": semana,
                                "tipo": tipo,
                                "categoría": categoria,
                                "descripción": concepto,
                                "importe": importe,
                            }
                        ]
                    )
                    df_actualizado = pd.concat([df, nuevo], ignore_index=True)
                    df_actualizado.to_excel(ARCHIVO_DATOS, index=False)
                    generar_flujo(df_actualizado)
                    st.session_state.concepto_seleccionado = None
                    st.session_state.categoria_seleccionada = None
                    st.success("Movimiento agregado correctamente y flujo actualizado.")

    # ---------------- ELIMINAR MOVIMIENTO ----------------
    elif seccion == "eliminar":
        with st.container(border=True):
            df_eliminar = pd.read_excel(ARCHIVO_DATOS)
            df_eliminar["inicio_semana"] = pd.to_datetime(
                df_eliminar["inicio_semana"], dayfirst=True
            )
            df_eliminar["importe"] = pd.to_numeric(
                df_eliminar["importe"], errors="coerce"
            ).fillna(0)
            df_eliminar["movimiento"] = (
                df_eliminar["semana"].astype(str)
                + " | "
                + df_eliminar["tipo"].astype(str)
                + " | "
                + df_eliminar["categoría"].astype(str)
                + " | "
                + df_eliminar["descripción"].astype(str)
                + " | $"
                + df_eliminar["importe"].map("{:,.2f}".format)
            )

            movimiento = st.selectbox(
                "Selecciona el movimiento que deseas eliminar",
                [""] + df_eliminar["movimiento"].tolist(),
            )
            confirmar = st.checkbox("Confirmo que quiero eliminar este movimiento")

            if st.button("Eliminar movimiento"):
                if movimiento == "":
                    st.error("Debes seleccionar un movimiento.")
                elif not confirmar:
                    st.error("Debes confirmar antes de eliminar.")
                else:
                    indice = df_eliminar[df_eliminar["movimiento"] == movimiento].index[
                        0
                    ]
                    df_eliminar = df_eliminar.drop(index=indice)
                    df_eliminar = df_eliminar.drop(columns=["movimiento"])
                    df_eliminar.to_excel(ARCHIVO_DATOS, index=False)
                    generar_flujo(df_eliminar)
                    st.success(
                        "Movimiento eliminado correctamente y flujo actualizado."
                    )


# ============================================================
#  MÓDULO 2: PAGOS PENDIENTES
# ============================================================
elif st.session_state.modulo_activo == "pagos":
    seccion_p = st.session_state.seccion_pagos

    st.markdown(
        f"""
    <div class="encabezado-modulo">
        <div class="eyebrow">Pagos Pendientes</div>
        <h1>{TITULOS_SECCION_PAGOS[seccion_p]}</h1>
    </div>
    <div class="regla-marca"></div>
    """,
        unsafe_allow_html=True,
    )

    dfp = cargar_pagos()

    # ---------------- AGREGAR DEUDA ----------------
    if seccion_p == "agregar":
        with st.container(border=True):
            conceptos_existentes = obtener_conceptos(df)

            concepto_deuda = st.selectbox(
                "Selecciona el concepto ya registrado (proveedor / concepto)",
                [""] + conceptos_existentes,
            )

            categoria_deuda = ""
            tipo_deuda = ""
            if concepto_deuda != "":
                datos_concepto = df[
                    df["descripción"].astype(str).str.strip() == concepto_deuda
                ].iloc[0]
                categoria_deuda = datos_concepto["categoría"]
                tipo_deuda = datos_concepto["tipo"]
                st.text_input("Categoría", value=categoria_deuda, disabled=True)
                st.text_input("Tipo", value=tipo_deuda, disabled=True)

            col1, col2 = st.columns(2)
            with col1:
                inicio_deuda = st.date_input(
                    "Inicio de semana a la que corresponde la deuda",
                    key="deuda_inicio_semana",
                    on_change=actualizar_fin_deuda,
                )
            with col2:
                fin_deuda = st.date_input(
                    "Fin de semana a la que corresponde la deuda",
                    key="deuda_fin_semana",
                )

            semana_deuda = formato_semana(inicio_deuda, fin_deuda)
            st.info(f"Semana de la deuda: {semana_deuda}")

            importe_deuda = st.number_input(
                "Importe pendiente", min_value=0.0, step=100.0, key="importe_deuda"
            )

            if st.button("Registrar deuda"):
                if concepto_deuda == "":
                    st.error("Debes seleccionar un concepto.")
                elif importe_deuda <= 0:
                    st.error("El importe debe ser mayor a 0.")
                elif fin_deuda < inicio_deuda:
                    st.error("La fecha final no puede ser anterior a la fecha inicial.")
                else:
                    nueva_deuda = pd.DataFrame(
                        [
                            {
                                "concepto": concepto_deuda,
                                "categoría": categoria_deuda,
                                "tipo": tipo_deuda,
                                "importe": importe_deuda,
                                "semana": semana_deuda,
                                "inicio_semana": pd.to_datetime(inicio_deuda),
                                "fecha_registro": pd.to_datetime(date.today()),
                                "estatus": "Pendiente",
                                "fecha_pago": "",
                            }
                        ]
                    )
                    dfp_actualizado = pd.concat([dfp, nueva_deuda], ignore_index=True)
                    guardar_pagos(dfp_actualizado)
                    st.success("Deuda registrada correctamente como pago pendiente.")

        st.write("")
        pendientes_actuales = dfp[dfp["estatus"] == "Pendiente"]
        if not pendientes_actuales.empty:
            with st.container(border=True):
                st.caption("Pagos pendientes actuales (por semana)")
                st.dataframe(
                    pendientes_actuales[
                        [
                            "concepto",
                            "categoría",
                            "tipo",
                            "importe",
                            "semana",
                            "fecha_registro",
                        ]
                    ],
                    use_container_width=True,
                )

    # ---------------- PAGAR UN PAGO PENDIENTE ----------------
    elif seccion_p == "pagar":
        with st.container(border=True):
            pendientes = dfp[dfp["estatus"] == "Pendiente"].copy()
            pendientes["importe"] = pd.to_numeric(
                pendientes["importe"], errors="coerce"
            ).fillna(0)

            if pendientes.empty:
                st.info("No hay pagos pendientes registrados.")
            else:
                # Se agrupa por concepto: el proveedor puede tener deuda
                # acumulada de varias semanas distintas.
                totales_por_concepto = (
                    pendientes.groupby("concepto")["importe"].sum().sort_index()
                )
                opciones_concepto = totales_por_concepto.index.tolist()

                concepto_pago = st.selectbox(
                    "Selecciona el concepto a pagar", [""] + opciones_concepto
                )

                total_pendiente = 0.0
                if concepto_pago != "":
                    total_pendiente = float(totales_por_concepto[concepto_pago])
                    st.info(
                        f"Deuda pendiente total de {concepto_pago}: {moneda(total_pendiente)}"
                    )

                    detalle = pendientes[pendientes["concepto"] == concepto_pago][
                        ["semana", "importe", "fecha_registro"]
                    ].sort_values("fecha_registro")
                    st.caption("Detalle por semana:")
                    st.dataframe(detalle, use_container_width=True, hide_index=True)

                monto_pagar = st.number_input(
                    "Monto que se va a pagar",
                    min_value=0.0,
                    max_value=total_pendiente if total_pendiente > 0 else 0.0,
                    step=100.0,
                    key="monto_pagar",
                )

                confirmar_pago = st.checkbox("Confirmo que este monto fue liquidado")

                if st.button("Registrar pago"):
                    if concepto_pago == "":
                        st.error("Debes seleccionar un concepto.")
                    elif monto_pagar <= 0:
                        st.error("El monto a pagar debe ser mayor a 0.")
                    elif monto_pagar > total_pendiente:
                        st.error(
                            "El monto a pagar no puede ser mayor a la deuda pendiente."
                        )
                    elif not confirmar_pago:
                        st.error("Debes confirmar antes de registrar el pago.")
                    else:
                        # Se descuenta el monto de las deudas más antiguas
                        # primero, hasta agotar el pago realizado.
                        indices_concepto = (
                            pendientes[pendientes["concepto"] == concepto_pago]
                            .sort_values("fecha_registro")
                            .index.tolist()
                        )

                        datos_concepto_pago = pendientes[
                            pendientes["concepto"] == concepto_pago
                        ].iloc[0]

                        monto_restante = monto_pagar
                        for idx in indices_concepto:
                            if monto_restante <= 0:
                                break
                            saldo_fila = float(dfp.loc[idx, "importe"])
                            abono = min(saldo_fila, monto_restante)
                            nuevo_saldo = round(saldo_fila - abono, 2)
                            dfp.loc[idx, "importe"] = nuevo_saldo
                            monto_restante = round(monto_restante - abono, 2)
                            if nuevo_saldo <= 0:
                                dfp.loc[idx, "importe"] = 0
                                dfp.loc[idx, "estatus"] = "Pagado"
                                dfp.loc[idx, "fecha_pago"] = pd.to_datetime(
                                    date.today()
                                )

                        guardar_pagos(dfp)

                        # Se refleja el pago como egreso real en el flujo de
                        # efectivo, usando la semana actual (la del pago).
                        inicio_pago = date.today()
                        fin_pago = inicio_pago + timedelta(days=5)
                        semana_pago = formato_semana(inicio_pago, fin_pago)

                        nuevo_movimiento = pd.DataFrame(
                            [
                                {
                                    "inicio_semana": pd.to_datetime(inicio_pago),
                                    "semana": semana_pago,
                                    "tipo": datos_concepto_pago["tipo"],
                                    "categoría": datos_concepto_pago["categoría"],
                                    "descripción": concepto_pago,
                                    "importe": monto_pagar,
                                }
                            ]
                        )
                        df_actualizado = pd.concat(
                            [df, nuevo_movimiento], ignore_index=True
                        )
                        df_actualizado.to_excel(ARCHIVO_DATOS, index=False)
                        generar_flujo(df_actualizado)

                        restante = round(total_pendiente - monto_pagar, 2)
                        if restante <= 0:
                            st.success(
                                f"Pago registrado. La deuda de {concepto_pago} quedó liquidada por completo."
                            )
                        else:
                            st.success(
                                f"Pago registrado. Queda un saldo pendiente de {moneda(restante)} en {concepto_pago}."
                            )

        st.write("")
        pagados = dfp[dfp["estatus"] == "Pagado"]
        if not pagados.empty:
            with st.container(border=True):
                st.caption("Historial de pagos liquidados")
                st.dataframe(
                    pagados[
                        [
                            "concepto",
                            "categoría",
                            "importe",
                            "semana",
                            "fecha_registro",
                            "fecha_pago",
                        ]
                    ],
                    use_container_width=True,
                )

    # ---------------- OPTIMIZACIÓN DE PAGOS ----------------
    # ---------------- OPTIMIZACIÓN DE PAGOS (reorganizado) ----------------
    elif seccion_p == "optimizar":
        deudas = obtener_deudas_pendientes(dfp)
        hoy = pd.Timestamp.today().normalize()
        df_semanas = df.copy()
        df_semanas["fin_semana"] = df_semanas["inicio_semana"] + pd.Timedelta(days=6)
        df_completas = df_semanas.loc[df_semanas["fin_semana"] < hoy]
        semanas_disponibles = df_completas["inicio_semana"].nunique()
        max_ventana = max(1, semanas_disponibles)


        if deudas.empty:
            st.info(
                "No hay pagos pendientes registrados para optimizar. "
                "Registra deudas en 'Agregar deuda' primero."
            )
        else:
            # ── 1. RECOPILATORIO DE DEUDA ─────────────────────────────
            with st.container(border=True):
                st.caption("Deudas pendientes consideradas (D_i), sumadas por concepto")
                st.caption(
                    "Ordenadas de menor a mayor: como el snowball paga primero las "
                    "deudas más chicas, así se ve de inmediato cuáles se liquidan antes."
                )
                st.dataframe(
                    deudas.rename("Importe pendiente ($)"), use_container_width=True
                )

            # ── 2. OPTIMIZADOR (MODELO A – PARETO) ─────────────────────
            with st.container(border=True):
                st.subheader("Optimización de pagos — Frontera de Pareto x Media Movil Ponderada")
                st.caption(
                    "Compara distintos montos posibles de pago total y elige el más "
                    "cercano al punto ideal (máxima caja, mínima deuda). Reparte ese "
                    "monto con snowball. La proyección de ingresos y egresos usa una "
                    "media móvil ponderada de las últimas 8 semanas de historial "
                    "(las semanas más recientes pesan más)."
                )

                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    ventana = st.number_input(
                        "Ventana del Promedio Movil (semanas)",
                        min_value = 1,
                        max_value = max_ventana,
                        value = min(8,max_ventana),
                        step = 1,
                        help = f"Máximo de semanas completas en el historial {max_ventana}"
                    )
                with col2:
                    horizonte = st.number_input(
                        "Horizonte de proyección (semanas, T)",
                        min_value=1,
                        max_value=52,
                        value=4,
                        step=1,
                    )
                with col3:
                    umbral_verde = st.number_input(
                        "Umbral verde (colchón cómodo)",
                        min_value=0.0,
                        value=200_000.0,
                        step=500.0,
                    )
                with col4:
                    umbral_amarillo = st.number_input(
                        "Umbral amarillo (precaución)",
                        min_value=0.0,
                        value=50_000.0,
                        step=500.0,
                    )
                with col5:
                    umbral_rojo = st.number_input(
                        "Umbral rojo (piso de emergencia)",
                        min_value=0.0,
                        value=10_000.0,
                        step=500.0,
                    )


                saldo_actual = calcular_saldo_actual(df)
                I_proy, E_proy = obtener_proyeccion_ponderada(df, horizonte, ventana)
                st.metric("Saldo actual (C0)", moneda(saldo_actual))
                st.caption(
                    f"Proyección media móvil ponderada ({ventana} sem.): {moneda(I_proy[0])}/semana "
                    f"de ingreso, {moneda(E_proy[0])}/semana de egreso (las semanas más "
                    f"recientes pesan más), aplicado a las próximas {horizonte} semanas."
                )

                conceptos = deudas.index.tolist()
                D0 = deudas.values.tolist()

                pago_por_concepto, diag, exito = modelo_a_pareto(
                    D0, I_proy, E_proy, saldo_actual, umbral_rojo, horizonte
                )

                if not exito:
                    st.error(
                        "El modelo no encontró una solución factible: con el saldo "
                        "actual y la proyección a este horizonte, ni pagando $0 de "
                        "deuda se respeta el piso rojo. Revisa el saldo actual, el "
                        "horizonte o el umbral rojo."
                    )
                else:
                    pago_total = diag["pago_total"]
                    caja_resultante = diag["caja_resultante"]
                    caja_proyectada = diag["caja_proyectada"]
                    deuda_resultante = diag["deuda_resultante"]

                    zona_result, _, delta_result = clasificar_caja(
                        caja_resultante, umbral_verde, umbral_amarillo, umbral_rojo
                    )
                    zona_proy, _, delta_proy = clasificar_caja(
                        caja_proyectada, umbral_verde, umbral_amarillo, umbral_rojo
                    )

                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Pago total recomendado", moneda(pago_total))
                    c2.metric("Caja resultante (solo el pago)", moneda(caja_resultante))
                    c3.metric(f"Caja  r. proyectada (a {horizonte} sem.)", moneda(caja_proyectada))
                    c4.metric("Deuda restante", moneda(deuda_resultante))

                    st.markdown(
                        f"**Semáforo caja resultante (justo después de pagar): {zona_result}** "
                        f"({'faltan' if delta_result < 0 else 'sobran'} "
                        f"{moneda(abs(delta_result))} respecto al umbral de esa zona)"
                    )
                    st.markdown(
                        f"**Semáforo caja proyectada (a {horizonte} semanas): {zona_proy}** "
                        f"({'faltan' if delta_proy < 0 else 'sobran'} "
                        f"{moneda(abs(delta_proy))} respecto al umbral de esa zona)"
                    )

                    tabla = construir_tabla_lista_pagos(pago_por_concepto, conceptos)
                    st.dataframe(tabla, use_container_width=True, hide_index=True)

                    # Excel de recomendaciones
                    with pd.ExcelWriter(ARCHIVO_OPTIMIZACION, engine="openpyxl") as writer:
                        tabla.to_excel(writer, sheet_name="Optimización", index=False)
                        ws = writer.sheets["Optimización"]
                        aplicar_formato_tabla(ws)
                        for row in ws.iter_rows(min_row=2):
                            for cell in row:
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = "#,##0.00"

                    with open(ARCHIVO_OPTIMIZACION, "rb") as archivo:
                        st.download_button(
                            "Descargar recomendaciones en Excel",
                            data=archivo,
                            file_name=ARCHIVO_OPTIMIZACION,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="descarga_optimizacion",
                        )

# ============================================================
#  PANTALLA DE BIENVENIDA (cuando no hay ningún módulo abierto)
# ============================================================
else:
    st.markdown(
        f"""
    <div class="bienvenida-hero">
        <img src="{LOGO_BASE64}" class="logo-hero" alt="FV Procesados">
    </div>
    """,
        unsafe_allow_html=True,
    )
